# -*- coding: utf-8 -*-
import openpyxl, json

FN = 'Strategia SEO OCMD - kondycjonowanie-wody.pl (1) (1).xlsx'
wb = openpyxl.load_workbook(FN, data_only=True)

def val(c):
    return '' if c is None else str(c).strip()

# --- Optymalizacja: authoritative SEO fields ---
ws = wb['🛠 Optymalizacja']
# header row is R2: cols -> # | 1 2 3 4 5 6 | Adresy URL | Nazwa widoku | Role | Title | H1 | Meta Description | Breadcrumbs
opt = []
for r in range(3, ws.max_row + 1):
    name_levels = [val(ws.cell(r, c).value) for c in range(1, 8)]  # cols A..G (# + levels 1-6)
    label = next((x for x in name_levels if x), '')
    url = val(ws.cell(r, 8).value)
    view = val(ws.cell(r, 9).value)
    role = val(ws.cell(r, 10).value)
    title = val(ws.cell(r, 11).value)
    h1 = val(ws.cell(r, 12).value)
    meta = val(ws.cell(r, 13).value)
    crumbs = val(ws.cell(r, 14).value)
    # depth = index of the non-empty level (0-based among the 6 level cols B..G)
    depth = 0
    for i in range(1, 7):
        if val(ws.cell(r, i + 1).value):
            depth = i - 1
            break
    if not url and not title:
        continue
    opt.append({
        'label': label, 'url': url, 'view': view, 'role': role,
        'title': title, 'h1': h1, 'meta': meta, 'crumbs': crumbs, 'depth': depth,
    })

with open('_seo.json', 'w', encoding='utf-8') as f:
    json.dump(opt, f, ensure_ascii=False, indent=1)

print(f'Optymalizacja rows: {len(opt)}')
for p in opt:
    print(f"  d{p['depth']} | {p['url']}")
